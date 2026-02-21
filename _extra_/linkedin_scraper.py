# pyre-ignore-all-errors
"""
LinkedIn Post Analytics Scraper  v3.0  –  Export Button Edition
----------------------------------------------------------------
Per ogni post nell'Excel di input:
  1. Naviga alla pagina /analytics/post-summary/urn:li:activity:XXX/
  2. Clicca il pulsante "Esporta"
  3. Intercetta il file .xlsx scaricato da LinkedIn
  4. Legge i due fogli: RENDIMENTO + PRINCIPALI DATI DEMOGRAFICI
  5. Estrae anche il testo del post dalla pagina
  6. Consolida tutto in un unico Excel di output

Struttura del file esportato da LinkedIn (osservata sul file reale):
  Foglio RENDIMENTO (coppie label→valore):
    URL post, Data di pubblicazione, Ora di pubblicazione del post,
    Impressioni, Utenti raggiunti, Visitatori del profilo da questo post,
    Follower acquisiti da questo post, Reazioni, Commenti, Diffusioni post,
    Salvataggi, Invii su LinkedIn,
    Qualifica principale (reazioni), Località principale (reazioni), Settore principale (reazioni),
    Qualifica principale (commenti), Località principale (commenti), Settore principale (commenti)

  Foglio PRINCIPALI DATI DEMOGRAFICI (tabella Categoria | Valore | %):
    Dimensioni azienda, Qualifica, Località, Azienda, Settore, Anzianità

UTILIZZO:
  python linkedin_scraper.py --input posts.xlsx --output stats_output.xlsx

REQUISITI:
  pip install playwright openpyxl pandas
  playwright install chromium
"""

import argparse
import sys
import time
import re
import tempfile
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
except ImportError:
    print("ERRORE: playwright non installato.")
    print("Esegui: pip install playwright && playwright install chromium")
    sys.exit(1)


# ─── utility ──────────────────────────────────────────────────────────────────

def read_input_excel(path: str) -> list:
    df = pd.read_excel(path, dtype=str)
    url_col = None
    for col in df.columns:
        if any(kw in col.lower() for kw in ("url", "link", "post", "href")):
            url_col = col
            break
    if url_col is None:
        url_col = df.columns[0]
    urls = df[url_col].dropna().str.strip().tolist()
    urls = [u for u in urls if u.startswith("http")]
    print(f"  Trovati {len(urls)} URL nella colonna '{url_col}'")
    return urls


def analytics_url(post_url: str) -> str:
    clean = post_url.strip().rstrip("/")
    m = re.search(r"(urn:li:(?:activity|ugcPost):\d+)", clean)
    if m:
        return f"https://www.linkedin.com/analytics/post-summary/{m.group(1)}/"
    m = re.search(r"activity-(\d{10,})", clean)
    if m:
        return f"https://www.linkedin.com/analytics/post-summary/urn:li:activity:{m.group(1)}/"
    return clean


def pct_str(val) -> str:
    """Converte 0.275 → '27.5%'"""
    try:
        return f"{float(val)*100:.1f}%"
    except (ValueError, TypeError):
        return str(val) if val else ""


# ─── parsing del file Excel esportato da LinkedIn ─────────────────────────────

def parse_linkedin_export(xlsx_path: str) -> dict:
    """
    Legge il file xlsx esportato da LinkedIn e restituisce un dict con
    tutti i campi estratti dai fogli RENDIMENTO e PRINCIPALI DATI DEMOGRAFICI.
    """
    result = {}

    try:
        xl = pd.ExcelFile(xlsx_path)
        sheet_names_lower = {s.lower(): s for s in xl.sheet_names}

        # ── Foglio RENDIMENTO ────────────────────────────────────────────────
        rendimento_key = next(
            (v for k, v in sheet_names_lower.items() if "rendimento" in k or "performance" in k),
            xl.sheet_names[0]
        )
        demo_key = next(
            (v for k, v in sheet_names_lower.items() if "demograf" in k or "demographic" in k),
            xl.sheet_names[1] if len(xl.sheet_names) > 1 else None
        )
        # Leggi tutti i fogli necessari mentre il file è aperto, poi chiudi subito
        df_r = pd.read_excel(xlsx_path, sheet_name=rendimento_key, header=None, dtype=str)
        df_d = pd.read_excel(xlsx_path, sheet_name=demo_key, header=0, dtype=str) if demo_key else None
        xl.close()  # chiude il file handle – fondamentale su Windows

        # Costruiamo un dizionario label→valore dalle coppie colonna 0/colonna 1
        kv = {}
        for _, row in df_r.iterrows():
            label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            value = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            if label and label != "nan":
                kv[label.lower()] = value

        def get(keys: list) -> str:
            for k in keys:
                for kv_key, val in kv.items():
                    if k.lower() in kv_key:
                        return val if val != "nan" else ""
            return ""

        result["post_url_export"]   = get(["url post", "post url"])
        result["post_date"]         = get(["data di pubblicazione", "publication date"])
        result["post_time"]         = get(["ora di pubblicazione", "publication time"])
        result["impressions"]       = get(["impressioni", "impressions"])
        result["unique_views"]      = get(["utenti raggiunti", "unique viewers", "reach"])
        result["profile_visits"]    = get(["visitatori del profilo", "profile visitors"])
        result["followers_gained"]  = get(["follower acquisiti", "followers gained"])
        result["reactions"]         = get(["reazioni", "reactions"])
        result["comments"]          = get(["commenti", "comments"])
        result["reposts"]           = get(["diffusioni", "reposts", "reshares"])
        result["saves"]             = get(["salvataggi", "saves"])
        result["sends"]             = get(["invii", "sends"])

        # Top reazioni
        # Il foglio ha sezioni separate per reazioni e commenti
        # Individuiamo le righe per sezione usando l'indice
        reaz_section = False
        comm_section = False
        reaz_rows = []
        comm_rows = []
        for _, row in df_r.iterrows():
            label = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ""
            if "reazioni in evidenza" in label:
                reaz_section = True
                comm_section = False
                continue
            if "commenti in evidenza" in label:
                comm_section = True
                reaz_section = False
                continue
            if reaz_section and label:
                reaz_rows.append((str(row.iloc[0]).strip(), str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""))
            if comm_section and label:
                comm_rows.append((str(row.iloc[0]).strip(), str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""))

        def section_val(rows, keyword):
            for label, val in rows:
                if keyword.lower() in label.lower() and val and val != "nan":
                    return val
            return ""

        result["reactions_top_role"]     = section_val(reaz_rows, "qualifica")
        result["reactions_top_location"] = section_val(reaz_rows, "localit")
        result["reactions_top_industry"] = section_val(reaz_rows, "settore")
        result["comments_top_role"]      = section_val(comm_rows, "qualifica")
        result["comments_top_location"]  = section_val(comm_rows, "localit")
        result["comments_top_industry"]  = section_val(comm_rows, "settore")

        # ── Foglio PRINCIPALI DATI DEMOGRAFICI ──────────────────────────────
        if df_d is not None:
            # Colonne: Categoria | Valore | %
            col_cat = df_d.columns[0]
            col_val = df_d.columns[1]
            col_pct = df_d.columns[2]

            def top_demo(category_kw: str, n: int = 3) -> str:
                mask = df_d[col_cat].str.lower().str.contains(category_kw, na=False)
                sub = df_d[mask].head(n)
                parts = []
                for _, row in sub.iterrows():
                    val = str(row[col_val]).strip()
                    pct = pct_str(row[col_pct])
                    if val and val != "nan":
                        parts.append(f"{val} ({pct})")
                return " | ".join(parts)

            result["demo_seniority"]    = top_demo("anzianit")
            result["demo_role"]         = top_demo("qualifica")
            result["demo_industry"]     = top_demo("settore")
            result["demo_company_size"] = top_demo("dimensioni azienda")
            result["demo_location"]     = top_demo("localit")
            result["demo_company"]      = top_demo("^azienda")

    except Exception as exc:
        result["parse_error"] = str(exc)[:200]
        print(f"    ✗ Errore parsing export: {exc}")

    return result


# ─── estrazione testo post dalla pagina ───────────────────────────────────────

def extract_post_text(page_text: str) -> str:
    """
    Estrae il testo del post dalla pagina analytics (prima di 'Scoperta').
    Usata come fallback se la visita al post originale fallisce.
    """
    parts = re.split(r"\nScoperta\n", page_text, maxsplit=1)
    if len(parts) < 2:
        return ""
    block = parts[0]
    m = re.search(r".+ha pubblicato questo post\s*[•·]\s*\S+\n?", block)
    if m:
        block = block[m.end():]
    else:
        lines = block.split("\n")
        start = next((i for i, l in enumerate(lines) if len(l.strip()) > 30 and i > 5), 0)
        block = "\n".join(lines[start:])
    block = re.sub(r"\nhashtag\n", " ", block)
    block = re.sub(r"[…\.]{3}visualizza altro\s*$", "", block)
    half = len(block) // 2
    if block[:50].strip() and block[half:half+50].strip().startswith(block[:30].strip()):
        block = block[:half]
    return block.strip()


def extract_post_text_from_post_page(page_text: str) -> str:
    """
    Estrae il testo COMPLETO del post dalla pagina del post originale
    (dopo aver cliccato 'visualizza altro').
    La pagina del post ha struttura diversa da quella analytics:
    il testo si trova tra l'header autore e la sezione commenti/reazioni.
    """
    # Rimuovi rumore di navigazione iniziale (barra nav, notifiche, ecc.)
    # Cerca l'inizio del post: di solito dopo "• Xh" o "• Xm" (tempo relativo)
    # oppure dopo una data tipo "17 nov"
    m = re.search(
        r"ha pubblicato questo post\s*[•·][^\n]*\n"   # header autore
        r"(?:[^\n]*\n){0,3}",                           # eventuali righe accessorie
        page_text
    )
    if m:
        block = page_text[m.end():]
    else:
        # Fallback: salta le prime 15 righe di nav
        lines = page_text.split("\n")
        start = next((i for i, l in enumerate(lines) if len(l.strip()) > 40 and i > 10), 0)
        block = "\n".join(lines[start:])

    # Il testo del post finisce quando cominciano i commenti / reazioni
    # Indicatori di fine post:
    end_markers = [
        r"\nReazioni\n",
        r"\nCommenti\n",
        r"\nAggiungi un commento",
        r"\nRispondi",
        r"\nMi piace\s*\n",
        r"\nCondividi\n",
        r"\nInvia\n",
        r"\nPost correlati",
        r"\nPotrebbe interessarti",
        r"\nAltri post di",
        r"\n\d+ reazioni?\n",
        r"\n\d+ commenti?\n",
    ]
    earliest = len(block)
    for marker in end_markers:
        m2 = re.search(marker, block, re.IGNORECASE)
        if m2 and m2.start() < earliest:
            earliest = m2.start()
    block = block[:earliest]

    # Pulizia
    block = re.sub(r"\nhashtag\n", " ", block)
    block = re.sub(r"[…\.]{3}visualizza altro\s*", "", block)
    block = block.strip()

    # Sanity check: se troppo corto, probabilmente qualcosa è andato storto
    if len(block) < 20:
        return ""

    return block


# ─── scraping principale ──────────────────────────────────────────────────────

def scrape_post(page, post_url: str, download_dir: Path) -> dict:
    result = {
        "post_url":              post_url,
        "post_text":             "",
        "post_date":             "",
        "post_time":             "",
        "impressions":           "",
        "unique_views":          "",
        "profile_visits":        "",
        "followers_gained":      "",
        "reactions":             "",
        "comments":              "",
        "reposts":               "",
        "saves":                 "",
        "sends":                 "",
        "reactions_top_role":    "",
        "reactions_top_location":"",
        "reactions_top_industry":"",
        "comments_top_role":     "",
        "comments_top_location": "",
        "comments_top_industry": "",
        "demo_seniority":        "",
        "demo_role":             "",
        "demo_industry":         "",
        "demo_company_size":     "",
        "demo_location":         "",
        "demo_company":          "",
        "scraped_at":            datetime.now().strftime("%Y-%m-%d %H:%M"),
        "error":                 "",
    }

    try:
        a_url = analytics_url(post_url)
        print(f"    → {a_url}")

        # ── 1. Naviga alla pagina analytics ───────────────────────────────
        page.goto(a_url, wait_until="domcontentloaded", timeout=30000)

        try:
            page.wait_for_selector(
                "button:has-text('Esporta'), button:has-text('Export'), "
                "[aria-label*='Esporta'], [aria-label*='Export']",
                timeout=15000
            )
        except PlaywrightTimeout:
            pass

        time.sleep(4)

        # Controllo redirect login
        if any(x in page.url for x in ("authwall", "/login", "checkpoint", "uas/authenticate")):
            result["error"] = "Redirect al login – sessione scaduta"
            print("    ✗ Redirect login")
            return result

        # ── 2. Visita il post originale e leggi il testo COMPLETO ──────────
        #    La pagina analytics tronca il testo con "visualizza altro".
        #    Visitiamo quindi direttamente l'URL del post, clicchiamo
        #    "visualizza altro" per espandere, e solo dopo torniamo all'analytics.
        try:
            page.goto(post_url, wait_until="domcontentloaded", timeout=20000)
            time.sleep(3)

            # Clicca tutti i pulsanti "visualizza altro" / "see more" presenti
            for see_more_sel in [
                "button.feed-shared-inline-show-more-text__see-more-less-toggle",
                "button:has-text('visualizza altro')",
                "button:has-text('…visualizza altro')",
                "button:has-text('see more')",
                "[aria-label*='visualizza altro']",
            ]:
                try:
                    btns = page.query_selector_all(see_more_sel)
                    for btn in btns:
                        if btn.is_visible():
                            btn.click()
                            time.sleep(0.5)
                except Exception:
                    pass

            time.sleep(1)
            post_page_text = page.inner_text("body")
            result["post_text"] = extract_post_text_from_post_page(post_page_text)
        except Exception as e_txt:
            print(f"    ⚠ Impossibile leggere testo dal post originale: {e_txt}")
            result["post_text"] = ""

        # ── 3. Torna alla pagina analytics, clicca "Esporta" ─────────────
        page.goto(a_url, wait_until="domcontentloaded", timeout=30000)
        try:
            page.wait_for_selector(
                "button:has-text('Esporta'), button:has-text('Export'), "
                "[aria-label*='Esporta'], [aria-label*='Export']",
                timeout=15000
            )
        except PlaywrightTimeout:
            pass
        time.sleep(3)

        page_text = page.inner_text("body")  # usato nel fallback stats

        # ── 4. Clicca "Esporta" e intercetta il download ──────────────────
        export_btn = None
        for selector in [
            "button:has-text('Esporta')",
            "button:has-text('Export')",
            "[aria-label*='Esporta']",
            "[aria-label*='Export']",
            "text=Esporta",
            "text=Export",
        ]:
            try:
                btn = page.query_selector(selector)
                if btn and btn.is_visible():
                    export_btn = btn
                    break
            except Exception:
                pass

        if not export_btn:
            result["error"] = "Pulsante Esporta non trovato"
            print("    ✗ Pulsante Esporta non trovato")
            # Fallback: prova a estrarre le stats dal testo della pagina
            _fill_stats_from_text(result, page_text)
            return result

        # Intercetta il download
        with page.expect_download(timeout=30000) as download_info:
            export_btn.click()

        download = download_info.value

        # Salva il file nella cartella temporanea
        file_name = download.suggested_filename or f"export_{int(time.time())}.xlsx"
        save_path = download_dir / file_name
        download.save_as(str(save_path))
        print(f"    ↓ Download: {file_name}")

        # ── 5. Parsa il file scaricato ─────────────────────────────────────
        export_data = parse_linkedin_export(str(save_path))
        result.update(export_data)

        # L'URL nel file export può differire (ugcPost vs activity); teniamo l'originale
        result["post_url"] = post_url

        # Pulizia file temporaneo
        try:
            save_path.unlink()
        except Exception:
            pass

        print(
            f"    ✓  impr={result['impressions'] or '—':>6}  "
            f"reach={result['unique_views'] or '—':>6}  "
            f"react={result['reactions'] or '—':>4}  "
            f"comm={result['comments'] or '—':>3}  "
            f"testo={len(result['post_text'])}ch"
        )

    except PlaywrightTimeout:
        result["error"] = "Timeout"
        print("    ✗ Timeout")
    except Exception as exc:
        result["error"] = str(exc)[:150]
        print(f"    ✗ Errore: {exc}")

    return result


def _fill_stats_from_text(result: dict, text: str):
    """Fallback: estrae statistiche dal testo della pagina se il download fallisce."""
    def after(label):
        m = re.search(rf"^{re.escape(label)}\n+([\d\.,]+)", text, re.MULTILINE)
        return re.sub(r"[.,](?=\d{{3}}(?!\d))", "", m.group(1)) if m else ""

    m = re.search(r"Scoperta\n+([\d\.,]+)", text)
    if m: result["impressions"] = re.sub(r"[.,](?=\d{3}(?!\d))", "", m.group(1))
    m = re.search(r"Impressioni\n+([\d\.,]+)", text)
    if m: result["unique_views"] = re.sub(r"[.,](?=\d{3}(?!\d))", "", m.group(1))
    result["reactions"] = after("Reazioni")
    result["comments"]  = after("Commenti")
    result["reposts"]   = after("Diffusioni post")
    result["saves"]     = after("Salvataggi")
    result["sends"]     = after("Invii su LinkedIn")


# ─── output Excel ─────────────────────────────────────────────────────────────

COLUMNS = [
    # Testo e date
    ("Testo del Post",               62),
    ("Data Pubblicazione",           18),
    ("Ora",                          10),
    # Metriche principali
    ("Impressioni",                  13),
    ("Utenti Raggiunti",             16),
    ("Reazioni",                     11),
    ("Commenti",                     11),
    ("Diffusioni",                   11),
    ("Salvataggi",                   11),
    ("Invii LinkedIn",               13),
    ("Visite Profilo",               13),
    ("Follower Acquisiti",           16),
    # Top reazioni
    ("React. – Qualifica",           26),
    ("React. – Città",               22),
    ("React. – Settore",             28),
    # Top commenti
    ("Comm. – Qualifica",            26),
    ("Comm. – Città",                22),
    ("Comm. – Settore",              28),
    # Demografici completi (top 3 per categoria)
    ("Demo: Anzianità",              40),
    ("Demo: Qualifiche",             40),
    ("Demo: Settori",                40),
    ("Demo: Dim. Azienda",           36),
    ("Demo: Città",                  36),
    ("Demo: Aziende",                40),
    # Meta
    ("Post URL",                     55),
    ("Estratto Il",                  18),
    ("Errore",                       35),
]

FIELD_MAP = [
    "post_text", "post_date", "post_time",
    "impressions", "unique_views",
    "reactions", "comments", "reposts", "saves", "sends",
    "profile_visits", "followers_gained",
    "reactions_top_role", "reactions_top_location", "reactions_top_industry",
    "comments_top_role", "comments_top_location", "comments_top_industry",
    "demo_seniority", "demo_role", "demo_industry",
    "demo_company_size", "demo_location", "demo_company",
    "post_url", "scraped_at", "error",
]

NUMERIC_COLS = {
    "Impressioni", "Utenti Raggiunti", "Reazioni", "Commenti",
    "Diffusioni", "Salvataggi", "Invii LinkedIn",
    "Visite Profilo", "Follower Acquisiti",
}


def save_to_excel(records: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "LinkedIn Stats"

    hdr_fill  = PatternFill("solid", start_color="0A66C2")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, bottom=thin, top=thin)

    # Raggruppa le colonne per sezione con colori header diversi
    section_colors = {
        "Testo del Post": "0A66C2",        "Data Pubblicazione": "0A66C2", "Ora": "0A66C2",
        "Impressioni": "1B6CA8",           "Utenti Raggiunti": "1B6CA8",
        "Reazioni": "1B6CA8",              "Commenti": "1B6CA8",
        "Diffusioni": "1B6CA8",            "Salvataggi": "1B6CA8",
        "Invii LinkedIn": "1B6CA8",        "Visite Profilo": "1B6CA8",
        "Follower Acquisiti": "1B6CA8",
        "React. – Qualifica": "2E7D32",    "React. – Città": "2E7D32",    "React. – Settore": "2E7D32",
        "Comm. – Qualifica": "1B5E20",     "Comm. – Città": "1B5E20",     "Comm. – Settore": "1B5E20",
        "Demo: Anzianità": "6A1B9A",       "Demo: Qualifiche": "6A1B9A",  "Demo: Settori": "6A1B9A",
        "Demo: Dim. Azienda": "6A1B9A",    "Demo: Città": "6A1B9A",       "Demo: Aziende": "6A1B9A",
        "Post URL": "37474F",              "Estratto Il": "37474F",        "Errore": "C62828",
    }

    for ci, (col_name, col_w) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=ci, value=col_name)
        color = section_colors.get(col_name, "0A66C2")
        c.font = hdr_font
        c.fill = PatternFill("solid", start_color=color)
        c.alignment = hdr_align
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = col_w

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    fill_even  = PatternFill("solid", start_color="EBF3FB")
    fill_odd   = PatternFill("solid", start_color="FFFFFF")
    data_font  = Font(name="Arial", size=10)
    wrap_al    = Alignment(wrap_text=True, vertical="top")
    center_al  = Alignment(horizontal="center", vertical="top")
    url_font   = Font(name="Arial", size=10, color="0A66C2", underline="single")

    for ri, record in enumerate(records, start=2):
        fill = fill_even if ri % 2 == 0 else fill_odd
        for ci, (field, (col_name, _)) in enumerate(zip(FIELD_MAP, COLUMNS), start=1):
            value = record.get(field, "")
            c = ws.cell(row=ri, column=ci, value=value)
            c.fill = fill
            c.border = border

            if col_name == "Testo del Post":
                c.font = data_font
                c.alignment = wrap_al
            elif col_name == "Post URL":
                c.font = url_font
                c.alignment = Alignment(vertical="top")
                if value:
                    try:
                        c.hyperlink = value
                    except Exception:
                        pass
            elif col_name in NUMERIC_COLS:
                c.font = data_font
                c.alignment = center_al
                try:
                    c.value = int(str(value).replace(".", "").replace(",", "")) if value else ""
                except (ValueError, TypeError):
                    c.value = value
            else:
                c.font = data_font
                c.alignment = Alignment(vertical="top", wrap_text=True)

        ws.row_dimensions[ri].height = 90

    # ── Foglio Riepilogo ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Riepilogo")
    ws2["A1"] = "LinkedIn Post Analytics – Riepilogo"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial", color="0A66C2")

    ok    = [r for r in records if not r.get("error")]
    errs  = [r for r in records if r.get("error")]
    with_text = [r for r in records if r.get("post_text")]

    stats_rows = [
        ("Post elaborati",              len(records)),
        ("Post completati con export",  len(ok)),
        ("Post con errori",             len(errs)),
        ("Post con testo estratto",     len(with_text)),
        ("Data estrazione",             datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for i, (label, val) in enumerate(stats_rows, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True, name="Arial")
        ws2.cell(row=i, column=2, value=val).font = Font(name="Arial")

    if errs:
        ws2.cell(row=9, column=1, value="Post con errori:").font = Font(bold=True, name="Arial", color="C62828")
        for j, r in enumerate(errs, start=10):
            ws2.cell(row=j, column=1, value=r["post_url"][:80]).font = Font(name="Arial", size=9)
            ws2.cell(row=j, column=2, value=r["error"]).font = Font(name="Arial", size=9, color="C62828")

    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 60

    wb.save(output_path)
    print(f"\n  ✅  {len(records)} record salvati → {output_path}")


# ─── main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="LinkedIn Post Analytics Scraper v3 – Export Button")
    parser.add_argument("--input",    "-i", required=True,
                        help="File Excel di input con gli URL dei post")
    parser.add_argument("--output",   "-o", default="linkedin_stats_output.xlsx",
                        help="File Excel di output (default: linkedin_stats_output.xlsx)")
    parser.add_argument("--delay",    "-d", type=float, default=4.0,
                        help="Secondi di attesa tra post (default: 4)")
    parser.add_argument("--headless", action="store_true",
                        help="Browser headless (sconsigliato: impedisce il login manuale)")
    args = parser.parse_args()

    if not Path(args.input).exists():
        print(f"ERRORE: File non trovato: {args.input}")
        sys.exit(1)

    print("\n╔══════════════════════════════════════════════════════════╗")
    print("║   LinkedIn Post Analytics Scraper  v3.0  –  Export      ║")
    print("╚══════════════════════════════════════════════════════════╝\n")

    print("📂 Lettura file di input …")
    urls = read_input_excel(args.input)
    if not urls:
        print("ERRORE: Nessun URL valido trovato.")
        sys.exit(1)

    records = []

    # Cartella temporanea per i download
    tmp_dir_obj = tempfile.TemporaryDirectory(ignore_cleanup_errors=True)
    tmp_dir = tmp_dir_obj.name
    try:
        download_dir = Path(tmp_dir)

        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=args.headless,
                args=["--start-maximized"],
                downloads_path=str(download_dir),
            )
            context = browser.new_context(
                viewport={"width": 1400, "height": 900},
                accept_downloads=True,
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
            )
            page = context.new_page()

            print("\n🔐 Apro LinkedIn … Effettua il login nel browser.")
            print("   (Attendo fino a 3 minuti)\n")
            page.goto("https://www.linkedin.com/login", wait_until="domcontentloaded")

            try:
                # Aspetta la barra di ricerca o la foto profilo (senza eval/unsafe-eval)
                page.wait_for_selector(
                    ".search-global-typeahead, .global-nav__me-photo, "
                    "[data-test-global-nav-search], #global-nav-search",
                    timeout=180000
                )
                print("   ✅  Login rilevato – avvio scraping …\n")
            except PlaywrightTimeout:
                print("   ⚠️  Timeout login, procedo comunque …\n")

            for i, url in enumerate(urls, start=1):
                print(f"[{i:>3}/{len(urls)}] {url[:90]}")
                record = scrape_post(page, url, download_dir)
                records.append(record)

                # Checkpoint ogni 10 post
                if i % 10 == 0:
                    tmp_out = args.output.replace(".xlsx", f"_checkpoint_{i}.xlsx")
                    save_to_excel(records, tmp_out)
                    print(f"  💾 Checkpoint: {tmp_out}")

                if i < len(urls):
                    time.sleep(args.delay)

            browser.close()

        # ── Salvataggio finale PRIMA di chiudere la cartella temporanea ───
        print("\n💾 Salvataggio finale …")
        save_to_excel(records, args.output)

        errors = [r for r in records if r["error"]]
        if errors:
            print(f"\n⚠️  {len(errors)} post con errori:")
            for r in errors:
                print(f"   • {r['post_url'][:70]}")
                print(f"     → {r['error']}")

    finally:
        # Pulizia cartella temporanea – su Windows i file xlsx scaricati
        # possono restare "locked"; ignoriamo l'errore se accade
        try:
            tmp_dir_obj.cleanup()
        except Exception:
            pass

    print("\nCompletato! 🎉")


if __name__ == "__main__":
    main()
